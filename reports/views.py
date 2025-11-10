from django.db.models import Sum, Count, F, ExpressionWrapper, DecimalField, Q, Avg as DjangoAvg
import django.db.models as models
from django.utils import timezone
from datetime import timedelta
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view, permission_classes
from rest_framework.parsers import MultiPartParser, FormParser
from products.models import Produit, Groupe
from journal.models import Mouvement
from .serializers import StockReportSerializer, MovementReportSerializer
from .models import ReorderAnalysis, ReorderConfiguration, ReorderMonitoring
from .services import ReorderCalculationService, ReorderActionService, ReorderCalculationServiceCompatible
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image
from django.http import HttpResponse, JsonResponse
import pandas as pd
import io
import logging
import json
import base64
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

logger = logging.getLogger(__name__)


class DashboardDataView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Ajouter des logs de débogage
        print("Récupération des données du tableau de bord")
        print(f"Utilisateur: {request.user.username}")
        
        try:
            # Obtenir la date d'aujourd'hui
            today = timezone.now().date()
            start_of_month = today.replace(day=1)

            # Statistiques de base
            total_products = Produit.objects.count()
            print(f"Nombre total de produits: {total_products}")
            
            total_groups = Groupe.objects.count()
            print(f"Nombre total de groupes: {total_groups}")
            
            low_stock_items = Produit.objects.filter(stock_initial__lt=F('seuil')).count()
            print(f"Produits en stock faible: {low_stock_items}")
            
            today_movements = Mouvement.objects.filter(date__date=today).count()

            # Calculer la valeur du stock (simulée pour l'instant)
            stock_value = Produit.objects.aggregate(
                total=Sum(ExpressionWrapper(F('stock_initial') * 10, output_field=DecimalField()))
            )['total'] or 0

            # Entrées et sorties du mois
            entries_this_month = Mouvement.objects.filter(
                date__date__gte=start_of_month,
                mouvement='Entrée'
            ).aggregate(total=Sum('quantite'))['total'] or 0

            exits_this_month = Mouvement.objects.filter(
                date__date__gte=start_of_month,
                mouvement='Sortie'
            ).aggregate(total=Sum('quantite'))['total'] or 0

            # Récupérer les mouvements récents
            recent_movements = Mouvement.objects.select_related('produit').order_by('-date')[:5]
            movements_data = [{
                'id': m.id,
                'produit': m.produit.nom,
                'mouvement': m.mouvement,
                'quantite': m.quantite,
                'demandeur': m.demandeur,
                'date': m.date
            } for m in recent_movements]

            # Récupérer les produits en stock faible
            low_stock_products = Produit.objects.filter(stock_initial__lt=F('seuil')).select_related('groupe')[:5]
            low_stock_data = [{
                'id': p.id,
                'code': p.code,
                'nom': p.nom,
                'stock_initial': p.stock_initial,
                'seuil': p.seuil,
                'groupe': p.groupe.nom if p.groupe else 'Non catégorisé'
            } for p in low_stock_products]

            # Données pour le graphique de stock
            groups = Groupe.objects.all()
            print(f"Groupes récupérés: {groups.count()}")
            stock_chart_data = []
            for g in groups:
                stock = Produit.objects.filter(groupe=g).aggregate(total=Sum('stock_initial'))['total'] or 0
                print(f"Groupe {g.nom}: {stock} unités en stock")
                stock_chart_data.append({
                    'groupe': g.nom,
                    'stock': stock
                })

            # Ajouter des données d'évolution des mouvements pour le graphique
            # Récupérer les 30 derniers jours
            end_date = timezone.now().date()
            start_date = end_date - timedelta(days=30)

            # Initialiser les données du graphique d'évolution
            stock_evolution_data = []

            # Générer des données pour chaque jour
            for i in range(31):  # 31 jours pour avoir un mois complet
                current_date = start_date + timedelta(days=i)
                date_str = current_date.strftime('%Y-%m-%d')
                
                # Récupérer les entrées et sorties pour cette date
                entries = Mouvement.objects.filter(
                    date__date=current_date,
                    mouvement='Entrée'
                ).aggregate(total=Sum('quantite'))['total'] or 0
                
                exits = Mouvement.objects.filter(
                    date__date=current_date,
                    mouvement='Sortie'
                ).aggregate(total=Sum('quantite'))['total'] or 0
                
                # Ajouter les données à la liste
                stock_evolution_data.append({
                    'date': date_str,
                    'entrees': entries,
                    'sorties': exits
                })

            response_data = {
                'stats': {
                    'totalProducts': total_products,
                    'totalGroups': total_groups,
                    'lowStockItems': low_stock_items,
                    'todayMovements': today_movements,
                    'entriesThisMonth': entries_this_month,
                    'exitsThisMonth': exits_this_month,
                },
                'recentMovements': movements_data,
                'lowStockAlerts': low_stock_data,
                'stockChartData': stock_evolution_data  # Utiliser les nouvelles données d'évolution
            }
            print("Données du tableau de bord prêtes à être envoyées")
            return Response(response_data)
        except Exception as e:
            print(f"ERREUR dans DashboardDataView: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class StockReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Ajouter des logs de débogage
        print("Récupération des données du rapport de stock")
        print(f"Utilisateur: {request.user.username}")
        
        try:
            # Filtres optionnels
            groupe_id = request.query_params.get('groupe')
            search = request.query_params.get('search')
            low_stock_only = request.query_params.get('low_stock') == 'true'
            date_start = request.query_params.get('dateStart')
            date_end = request.query_params.get('dateEnd')

            # Base de la requête
            queryset = Produit.objects.select_related('groupe').all()

            # Appliquer les filtres
            if groupe_id:
                queryset = queryset.filter(groupe_id=groupe_id)
            if search:
                queryset = queryset.filter(Q(nom__icontains=search) | Q(code__icontains=search))
            if low_stock_only:
                queryset = queryset.filter(stock_initial__lt=F('seuil'))

            # Vérifier si des produits ont été trouvés
            if queryset.count() == 0:
                print("Aucun produit trouvé avec les filtres appliqués")
                return Response({"message": "Aucun produit trouvé avec les filtres appliqués"}, status=status.HTTP_404_NOT_FOUND)
            
            # Initialiser la liste des produits
            products_data = []
            
            # Ajouter les produits à la liste
            for p in queryset:
                products_data.append({
                    'id': p.id, 
                    'code': p.code, 
                    'nom': p.nom, 
                    'stock_initial': p.stock_initial, 
                    'seuil': p.seuil, 
                    'groupe': {
                        'id': p.groupe.id,
                        'nom': p.groupe.nom,
                        'couleur': p.groupe.couleur
                        } if p.groupe else None
                })

            # Créer un dictionnaire avec les données attendues par le sérialiseur
            report_data = { 
                'date': timezone.now().date(),  # Date actuelle 
                'products': products_data, 
                'low_stock_count': queryset.filter(stock_initial__lt=F('seuil')).count() 
            } 

            print(f"Nombre de produits dans le rapport: {len(report_data['products'])}") 
            
            # Sérialiser les données
            serializer = StockReportSerializer(report_data) 
            return Response(serializer.data) 
        except Exception as e: 
            print(f"ERREUR dans StockReportView: {str(e)}") 
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class MovementReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Ajouter des logs de débogage
        print("Récupération des données du rapport de mouvements")
        print(f"Utilisateur: {request.user.username}")
        
        try:
            # Filtres optionnels
            start_date = request.query_params.get('start_date') or request.query_params.get('dateStart')
            end_date = request.query_params.get('end_date') or request.query_params.get('dateEnd')
            mouvement_type = request.query_params.get('type')
            produit_id = request.query_params.get('produit')
            groupe_id = request.query_params.get('groupe')
            zone = request.query_params.get('zone')

            # Base de la requête
            queryset = Mouvement.objects.select_related('produit', 'produit__groupe').all()

            # Appliquer les filtres
            if start_date:
                queryset = queryset.filter(date__date__gte=start_date)
            if end_date:
                queryset = queryset.filter(date__date__lte=end_date)
            if mouvement_type:
                queryset = queryset.filter(mouvement=mouvement_type)
            if produit_id:
                queryset = queryset.filter(produit_id=produit_id)
            if groupe_id:
                queryset = queryset.filter(produit__groupe_id=groupe_id)
            if zone:
                queryset = queryset.filter(produit__zone=zone)

            # Vérifier si des mouvements ont été trouvés
            if queryset.count() == 0:
                print("Aucun mouvement trouvé avec les filtres appliqués")

            # Préparer les données pour le sérialiseur
            movements_data = []
            for m in queryset:
                movements_data.append({
                    'id': m.id,
                    'produit': m.produit.nom,
                    'produit_id': m.produit.id,
                    'code': m.produit.code,
                    'mouvement': m.mouvement,
                    'quantite': m.quantite,
                    'stock_avant': m.stock_avant,
                    'stock_apres': m.stock_apres,
                    'demandeur': m.demandeur,
                    'date': m.date,
                    'observation': m.observation,
                    'groupe': m.produit.groupe.nom if m.produit.groupe else 'Non catégorisé'
                })

            # Calculer les statistiques
            stats = {
                'total_entries': queryset.filter(mouvement='Entrée').count(),
                'total_exits': queryset.filter(mouvement='Sortie').count(),
                'total_quantity': sum(m.quantite for m in queryset),
                'total_requests': queryset.filter(mouvement='Demandée').count(),
            }

            # Préparer les données de répartition par heure
            movements_by_hour = []
            for hour in range(24):
                # Filtrer les mouvements pour cette heure
                hour_movements = queryset.filter(date__hour=hour)
                movements_by_hour.append({
                    'hour': f"{hour}h",
                    'count': hour_movements.count(),
                    'entries': hour_movements.filter(mouvement='Entrée').count(),
                    'exits': hour_movements.filter(mouvement='Sortie').count(),
                })

            print(f"Nombre de mouvements dans le rapport: {len(movements_data)}")
            
            # Préparer les données pour le sérialiseur
            report_data = {
                'start_date': start_date,
                'end_date': end_date,
                'movements': movements_data,
                'movements_by_hour': movements_by_hour,  # Ajouter les données par heure
                'stats': stats
            }

            # Sérialiser les données
            serializer = MovementReportSerializer(report_data)
            return Response(serializer.data)
        except Exception as e:
            print(f"ERREUR dans MovementReportView: {str(e)}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ExportReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        report_type = request.query_params.get('type', 'stock')
        
        if report_type == 'stock':
            # Exporter le rapport de stock
            queryset = Produit.objects.select_related('groupe').all()
            
            # Créer un DataFrame pandas
            data = [{
                'Code': p.code,
                'Nom': p.nom,
                'Groupe': p.groupe.nom if p.groupe else 'Non catégorisé',
                'Stock': p.stock_initial,
                'Seuil': p.seuil,
                'Unité': p.unite,
                'Zone': p.zone,
                'Observation': p.observation
            } for p in queryset]
            
            df = pd.DataFrame(data)
            
        elif report_type == 'movements':
            # Exporter le rapport de mouvements
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            
            queryset = Mouvement.objects.select_related('produit').all()
            
            if start_date:
                queryset = queryset.filter(date__date__gte=start_date)
            if end_date:
                queryset = queryset.filter(date__date__lte=end_date)
                
            data = [{
                'Date': m.date,
                'Produit': m.produit.nom,
                'Code': m.produit.code,
                'Mouvement': m.mouvement,
                'Quantité': m.quantite,
                'Stock Après': m.stock_apres,
                'Demandeur': m.demandeur,
                'Observation': m.observation
            } for m in queryset]
            
            df = pd.DataFrame(data)
        else:
            return Response({'error': 'Type de rapport non valide'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Créer un buffer pour le fichier Excel
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Rapport', index=False)
            
            # Ajuster la largeur des colonnes
            worksheet = writer.sheets['Rapport']
            for i, col in enumerate(df.columns):
                column_width = max(df[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.set_column(i, i, column_width)
        
        buffer.seek(0)
        
        # Préparer la réponse avec le fichier Excel
        filename = f"rapport_{report_type}_{timezone.now().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


class ReorderRecommendationsView(APIView):
    """API pour les recommandations de point de commande"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        """Récupérer les recommandations de point de commande"""
        try:
            # Paramètres de filtrage
            status_filter = request.query_params.get('status', None)
            confidence_filter = request.query_params.get('confidence', None)
            depot_filter = request.query_params.get('depot', 'Principal')
            page = max(1, int(request.query_params.get('page', 1)))  # Correction: assurer page >= 1
            page_size = max(1, min(100, int(request.query_params.get('page_size', 20))))  # Correction: limiter page_size
            
            # Query de base
            queryset = ReorderAnalysis.objects.select_related('produit', 'produit__groupe').filter(
                depot=depot_filter
            )
            
            # Appliquer les filtres
            if status_filter:
                queryset = queryset.filter(status=status_filter)
            if confidence_filter:
                queryset = queryset.filter(confidence_level=confidence_filter)
            # Nouveau: filtre par groupe (id)
            groupe_filter = request.query_params.get('groupe', None)
            if groupe_filter and str(groupe_filter).isdigit():
                queryset = queryset.filter(produit__groupe_id=int(groupe_filter))
            
            # Pagination sécurisée
            total_count = queryset.count()
            
            if total_count == 0:
                return Response({
                    'success': True,
                    'data': {
                        'recommendations': [],
                        'pagination': {
                            'page': page,
                            'page_size': page_size,
                            'total_count': 0,
                            'total_pages': 0
                        },
                        'kpis': self._calculate_kpis(queryset.model.objects.filter(depot=depot_filter))
                    }
                })
            
            start_idx = (page - 1) * page_size
            end_idx = min(start_idx + page_size, total_count)  # Correction: éviter de dépasser total_count
            
            # Vérifier que start_idx est valide
            if start_idx >= total_count:
                page = 1
                start_idx = 0
                end_idx = min(page_size, total_count)
            
            analyses = queryset[start_idx:end_idx]
            
            # Sérialiser les données
            recommendations = []
            action_service = ReorderActionService()
            
            for analysis in analyses:
                actions = action_service.get_actions_for_analysis(analysis)
                
                recommendations.append({
                    'id': analysis.id,
                    'produit': {
                        'id': analysis.produit.id,
                        'code': analysis.produit.code,
                        'nom': analysis.produit.nom,
                        'groupe': (
                            {'id': analysis.produit.groupe.id, 'nom': analysis.produit.groupe.nom}
                            if analysis.produit.groupe else None
                        ),
                        'stock_actuel': analysis.produit.stock_initial,
                        'seuil': analysis.produit.seuil
                    },
                    'metrics': {
                        'avg_daily_demand': float(analysis.avg_daily_demand),
                        'sigma_daily_demand': float(analysis.sigma_daily_demand),
                        'lead_time_days': analysis.lead_time_days,
                        'safety_stock': float(analysis.safety_stock),
                        'reorder_point': float(analysis.reorder_point),
                        'predicted_30d': float(analysis.predicted_30d),
                        'days_of_coverage': float(analysis.days_of_coverage),
                        'suggested_qty': float(analysis.suggested_qty)
                    },
                    'status': analysis.status,
                    'confidence': {
                        'level': analysis.confidence_level,
                        'reason': analysis.confidence_reason
                    },
                    'calculation_info': {
                        'date': analysis.calculation_date,
                        'movements_count': analysis.movements_count,
                        'window_days_used': analysis.window_days_used,
                        'z_factor_used': float(analysis.z_factor_used)
                    },
                    'explanation': analysis.explanation,
                    'actions': actions
                })
            
            # KPIs globaux
            kpis = self._calculate_kpis(queryset.model.objects.filter(depot=depot_filter))
            
            return Response({
                'success': True,
                'data': {
                    'recommendations': recommendations,
                    'pagination': {
                        'page': page,
                        'page_size': page_size,
                        'total_count': total_count,
                        'total_pages': (total_count + page_size - 1) // page_size
                    },
                    'kpis': kpis
                }
            })
            
        except Exception as e:
            logger.error(f"Erreur récupération recommandations: {str(e)}")
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def _calculate_kpis(self, queryset):
        """Calculer les KPIs globaux avec gestion d'erreurs"""
        try:
            total = queryset.count()
            if total == 0:
                return {
                    'total_analyses': 0,
                    'urgent_count': 0,
                    'high_confidence': 0,
                    'avg_coverage_days': 0
                }
            
            urgent = queryset.filter(status='urgent').count()
            high_conf = queryset.filter(confidence_level='high').count()
            
            # Correction: utiliser Avg au lieu de Sum
            avg_coverage = queryset.aggregate(
                avg=models.Avg('days_of_coverage')
            )['avg'] or 0
            
            return {
                'total_analyses': total,
                'urgent_count': urgent,
                'high_confidence': high_conf,
                'avg_coverage_days': float(avg_coverage)
            }
        except Exception as e:
            logger.error(f"Erreur calcul KPIs: {str(e)}")
            return {
                'total_analyses': 0,
                'urgent_count': 0,
                'high_confidence': 0,
                'avg_coverage_days': 0
            }



@api_view(['POST'])
@permission_classes([IsAuthenticated])
def recalculate_product_analysis(request):
    """Recalcule l'analyse des produits pour les recommandations de réapprovisionnement (robuste)."""
    try:
        logger.info("Début recalc - utilisateur: %s", request.user.username)
        
        # Validation des données d'entrée
        data = request.data
        product_ids_raw = data.get('product_ids', None)
        
        logger.debug("Payload product_ids_raw: %s (type=%s)", product_ids_raw, type(product_ids_raw))
        
        # Normaliser product_ids en liste d'entiers valides
        product_ids = []
        
        if product_ids_raw is None:
            product_ids = []
        elif isinstance(product_ids_raw, list):
            # filtrer et convertir en int
            for item in product_ids_raw:
                try:
                    product_ids.append(int(item))
                except (ValueError, TypeError):
                    logger.warning("Ignorer product_id invalide dans la liste: %s", item)
        elif isinstance(product_ids_raw, str):
            # Peut être une chaîne JSON '["1","2"]' ou '1,2,3'
            try:
                parsed = json.loads(product_ids_raw)
                if isinstance(parsed, list):
                    for item in parsed:
                        try:
                            product_ids.append(int(item))
                        except (ValueError, TypeError):
                            logger.warning("Ignorer product_id invalide après json.loads: %s", item)
                else:
                    # fallback: split by comma
                    for part in product_ids_raw.split(','):
                        part = part.strip()
                        if part:
                            try:
                                product_ids.append(int(part))
                            except (ValueError, TypeError):
                                logger.warning("Ignorer product_id invalide après split: %s", part)
            except json.JSONDecodeError:
                # fallback: split by comma
                for part in product_ids_raw.split(','):
                    part = part.strip()
                    if part:
                        try:
                            product_ids.append(int(part))
                        except (ValueError, TypeError):
                            logger.warning("Ignorer product_id invalide après split fallback: %s", part)
        else:
            # cas non attendu (ex. dict). on essaye d'extraire des ids s'il existe une clé 'id'
            if isinstance(product_ids_raw, dict) and 'id' in product_ids_raw:
                try:
                    product_ids.append(int(product_ids_raw['id']))
                except (ValueError, TypeError):
                    pass
        
        logger.info("product_ids normalisés: %s", product_ids)
        
        # Vérifier la configuration
        config = ReorderConfiguration.objects.filter(is_active=True).first()
        if not config:
            return Response({
                'error': 'Configuration de réapprovisionnement manquante. Veuillez contacter l\'administrateur.'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        # Récupérer les produits
        if not product_ids:
            products = Produit.objects.all()
            logger.info("Recalcul pour tous les produits (%d)", products.count())
        else:
            products = Produit.objects.filter(id__in=product_ids)
            found_ids = list(products.values_list('id', flat=True))
            logger.info("Produits trouvés: %s (demandés: %s)", found_ids, product_ids)
            # signaler ids manquants mais ne pas planter
            missing_ids = [pid for pid in product_ids if pid not in found_ids]
            if missing_ids:
                logger.warning("Produits non trouvés: %s", missing_ids)
        
        if not products.exists():
            return Response({'error': 'Aucun produit valide trouvé'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Initialiser le service avec gestion d'erreurs
        try:
            service = ReorderCalculationService()
        except Exception as e:
            logger.error(f"Erreur initialisation service: {str(e)}")
            return Response({
                'error': f'Erreur d\'initialisation du service de calcul: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        success_count = 0
        error_count = 0
        errors = []
        results = []
        
        for produit in products:
            try:
                analysis = service.calculate_for_product(produit.id)  # Passer l'ID au lieu de l'objet
                if analysis:
                    success_count += 1
                    results.append({
                        'product_id': produit.id,
                        'product_name': produit.nom,
                        'status': analysis.status,
                        'reorder_point': float(analysis.reorder_point),
                        'suggested_qty': float(analysis.suggested_qty),
                        'success': True
                    })
                    logger.debug("Analyse OK pour %s", produit.code)
                else:
                    error_count += 1
                    errors.append({'product_id': produit.id, 'error': 'Aucune analyse générée'})
            except Exception as e:
                error_count += 1
                error_msg = str(e)
                errors.append({'product_id': product.id, 'error': error_msg})
                logger.exception("Erreur calcul produit %s: %s", product.code, error_msg)
        
        response_data = {
            'message': f'Recalcul terminé: {success_count} succès, {error_count} erreurs',
            'success_count': success_count,
            'error_count': error_count,
            'results': results,
            'errors': errors
        }
        
        status_code = status.HTTP_200_OK if error_count == 0 else (status.HTTP_207_MULTI_STATUS if success_count > 0 else status.HTTP_400_BAD_REQUEST)
        return Response(response_data, status=status_code)
        
    except Exception as e:
        logger.exception("Erreur générale lors du recalcul: %s", str(e))
        return Response({
            'success': False,
            'error': f"Erreur générale lors du recalcul: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ReorderMonitoringView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            logger.info("Récupération des données de monitoring")
            
            # Récupérer les enregistrements récents
            records = ReorderMonitoring.objects.all().order_by('-date')[:30]
            
            monitoring_data = []
            for record in records:
                # Calculer total_analyses à partir des champs existants
                total_analyses = record.urgent_count + record.monitor_count + record.ok_count
                
                monitoring_data.append({
                    'date': record.date.strftime('%Y-%m-%d'),
                    'urgent_count': record.urgent_count,
                    'monitor_count': record.monitor_count, 
                    'ok_count': record.ok_count,
                    'total_analyses': total_analyses,  # Calculé dynamiquement
                    'total_suggested_units': float(record.total_suggested_units),
                    'draft_orders_created': record.draft_orders_created,
                    'draft_orders_approved': record.draft_orders_approved,
                    'approval_rate': record.approval_rate,
                    'average_confidence': float(record.average_confidence) if record.average_confidence else 0,
                    'low_confidence_count': record.low_confidence_count,
                    'batch_execution_time': float(record.batch_execution_time) if record.batch_execution_time else 0,
                    'batch_success': record.batch_success
                })
            
            # Statistiques récentes
            latest_record = records.first() if records else None
            if latest_record:
                latest_total = latest_record.urgent_count + latest_record.monitor_count + latest_record.ok_count
            else:
                latest_total = 0
                
            return Response({
                'monitoring_data': monitoring_data,
                'latest_stats': {
                    'urgent_count': latest_record.urgent_count if latest_record else 0,
                    'monitor_count': latest_record.monitor_count if latest_record else 0,
                    'ok_count': latest_record.ok_count if latest_record else 0,
                    'total_analyses': latest_total,  # Calculé dynamiquement
                    'approval_rate': latest_record.approval_rate if latest_record else 0,
                    'average_confidence': float(latest_record.average_confidence) if latest_record and latest_record.average_confidence else 0
                }
            })
            
        except Exception as e:
            logger.error(f"Erreur récupération monitoring: {str(e)}")
            return Response(
                {'error': f'Erreur lors de la récupération des données de monitoring: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ExportXlsxAdvancedView(APIView):
    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [IsAuthenticated]
    
    def _convert_value_for_excel(self, value):
        """Convertit une valeur pour qu'elle soit compatible avec Excel"""
        if isinstance(value, (list, dict)):
            return str(value)  # Convertir en chaîne
        elif value is None:
            return ''
        elif isinstance(value, (int, float, str, bool)):
            return value
        else:
            return str(value)
    
    def _auto_adjust_columns(self, worksheet):
        """Ajuste automatiquement la largeur des colonnes"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Ajuster la largeur (minimum 10, maximum 50)
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _format_table_headers(self, worksheet, start_row, headers):
        """Formate les en-têtes de tableau"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=start_row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    def _create_info_sheet(self, ws, metadata):
        """Crée la feuille d'informations avec des détails sur le rapport"""
        ws['A1'] = "Rapport généré le"
        ws['B1'] = metadata.get('date', '')
        ws['A2'] = "Type de rapport"
        ws['B2'] = metadata.get('type', '')
        ws['A3'] = "Période"
        ws['B3'] = metadata.get('period', '')
        
        # Ajuster les colonnes
        self._auto_adjust_columns(ws)
        logger.info("Feuille Informations créée")
    
    def _create_enhanced_data_sheet(self, ws, tables_data, metadata):
        """Crée une feuille de données améliorée avec tableaux détaillés"""
        current_row = 1
        
        # Titre principal
        report_type = metadata.get('type', 'Rapport')
        ws.merge_cells(f'A{current_row}:F{current_row}')
        title_cell = ws[f'A{current_row}']
        title_cell.value = f"Données Détaillées - {report_type.title()}"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center')
        current_row += 3
        
        # Traitement selon le type de rapport
        if report_type == 'stock':
            current_row = self._add_stock_details_table(ws, current_row, metadata)
        elif report_type == 'movements':
            current_row = self._add_movements_details_table(ws, current_row, metadata)
        elif report_type == 'recommendations':
            current_row = self._add_recommendations_details_table(ws, current_row, metadata)
        
        # Ajout des tableaux existants
        for table in tables_data:
            current_row = self._add_table_to_sheet(ws, table, current_row)
            current_row += 2
        
        # Ajustement automatique des colonnes
        self._auto_adjust_columns(ws)
    
    def _add_table_to_sheet(self, ws, table, start_row):
        """Ajoute un tableau à la feuille de calcul"""
        # Titre du tableau avec formatage
        title_cell = ws.cell(row=start_row, column=1, value=table.get('title', ''))
        title_cell.font = Font(bold=True, size=14)
        start_row += 2
        
        # En-têtes avec formatage
        headers = table.get('headers', [])
        if headers:
            for col, header in enumerate(headers, 1):
                ws.cell(row=start_row, column=col, value=self._convert_value_for_excel(header))
            
            # Appliquer le formatage aux en-têtes
            self._format_table_headers(ws, start_row, headers)
            start_row += 1
        
        # Données du tableau
        data_rows = table.get('data', [])
        for data_row in data_rows:
            if isinstance(data_row, list):
                # Données sous forme de liste (format correct)
                for col, value in enumerate(data_row, 1):
                    converted_value = self._convert_value_for_excel(value)
                    ws.cell(row=start_row, column=col, value=converted_value)
            elif isinstance(data_row, dict):
                # Données sous forme d'objet (à convertir)
                col = 1
                for key, value in data_row.items():
                    converted_value = self._convert_value_for_excel(value)
                    ws.cell(row=start_row, column=col, value=converted_value)
                    col += 1
            start_row += 1
        
        return start_row
    
    def _add_stock_details_table(self, ws, start_row, metadata):
        """Ajoute le tableau détaillé des stocks"""
        try:
            from products.models import Produit
            
            # Titre du tableau
            ws.merge_cells(f'A{start_row}:F{start_row}')  # Changé de G à F
            title_cell = ws[f'A{start_row}']
            title_cell.value = "Détails des Stocks par Produit"
            title_cell.font = Font(bold=True, size=14)
            title_cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            start_row += 2
            
            # En-têtes - SUPPRESSION de 'Valeur Stock'
            headers = ['Code', 'Nom', 'Groupe', 'Stock Actuel', 'Seuil Min', 'Statut']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
            start_row += 1
            
            # Données des produits - SUPPRESSION de stock_value
            products = Produit.objects.select_related('groupe').all()
            for product in products:
                status = "Critique" if (product.stock_initial or 0) <= (product.seuil or 0) else "Normal"
                
                row_data = [
                    product.code or '',
                    product.nom or '',
                    product.groupe.nom if product.groupe else '',
                    product.stock_initial or 0,
                    product.seuil or 0,
                    status
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=start_row, column=col, value=self._convert_value_for_excel(value))
                    if status == "Critique" and col in [4, 6]:  # Stock et Statut (ajusté les indices)
                        cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                
                start_row += 1
            
            return start_row + 2
            
        except Exception as e:
            logger.error(f"Erreur lors de l'ajout du tableau de stock: {str(e)}")
            return start_row
    
    def _add_movements_details_table(self, ws, start_row, metadata):
        """Ajoute le tableau des top 10 mouvements"""
        try:
            from journal.models import Mouvement
            from django.db.models import Count, Sum
            
            # Titre du tableau
            ws.merge_cells(f'A{start_row}:F{start_row}')
            title_cell = ws[f'A{start_row}']
            title_cell.value = "Top 10 des Produits les Plus Actifs"
            title_cell.font = Font(bold=True, size=14)
            title_cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            start_row += 2
            
            # En-têtes
            headers = ['Rang', 'Code Produit', 'Nom Produit', 'Nb Mouvements', 'Quantité Totale', 'Dernière Activité']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
            start_row += 1
            
            # Récupération des données
            top_movements = Mouvement.objects.select_related('produit').values(
                'produit__code', 'produit__nom'
            ).annotate(
                movement_count=Count('id'),
                total_quantity=Sum('quantite')
            ).order_by('-movement_count')[:10]
            
            for rank, movement in enumerate(top_movements, 1):
                # Dernière activité
                last_movement = Mouvement.objects.filter(
                    produit__code=movement['produit__code']
                ).order_by('-date').first()
                
                last_activity = last_movement.date.strftime('%d/%m/%Y') if last_movement else 'N/A'
                
                row_data = [
                    rank,
                    movement['produit__code'] or '',
                    movement['produit__nom'] or '',
                    movement['movement_count'] or 0,
                    abs(movement['total_quantity'] or 0),
                    last_activity
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=start_row, column=col, value=self._convert_value_for_excel(value))
                    if rank <= 3:  # Top 3
                        cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
                
                start_row += 1
            
            return start_row + 2
            
        except Exception as e:
            logger.error(f"Erreur lors de l'ajout du tableau de mouvements: {str(e)}")
            return start_row
    
    def _add_recommendations_details_table(self, ws, start_row, metadata):
        """Ajoute le tableau des recommandations"""
        try:
            from reports.models import ReorderAnalysis
            
            # Titre du tableau
            ws.merge_cells(f'A{start_row}:G{start_row}')
            title_cell = ws[f'A{start_row}']
            title_cell.value = "Recommandations de Réapprovisionnement"
            title_cell.font = Font(bold=True, size=14)
            title_cell.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
            start_row += 2
            
            # En-têtes
            headers = ['Code', 'Nom Produit', 'Stock Actuel', 'Seuil', 'Qté Recommandée', 'Priorité', 'Date Analyse']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
            start_row += 1
            
            # Données des recommandations
            recommendations = ReorderAnalysis.objects.select_related('produit').filter(
                status__in=['urgent', 'monitor']
            ).order_by('status', '-calculation_date')
            
            for rec in recommendations:
                analysis_date = rec.calculation_date.strftime('%d/%m/%Y') if rec.calculation_date else 'N/A'
                
                row_data = [
                    rec.produit.code if rec.produit else '',
                    rec.produit.nom if rec.produit else '',
                    rec.produit.stock_initial or 0,
                    rec.produit.seuil or 0,
                    rec.suggested_qty or 0,
                    rec.status.upper() or 'OK',
                    analysis_date
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=start_row, column=col, value=self._convert_value_for_excel(value))
                    
                    # Coloration selon la priorité
                    if rec.status == 'urgent':
                        cell.fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
                    elif rec.status == 'monitor':
                        cell.fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
                
                start_row += 1
            
            return start_row + 2
            
        except Exception as e:
            logger.error(f"Erreur lors de l'ajout du tableau de recommandations: {str(e)}")
            return start_row
    
    def _create_charts_sheet(self, ws, chart_images):
        """Crée la feuille des graphiques"""
        chart_row = 1
        chart_count = 0
        
        for chart_info in chart_images:
            try:
                # Validation des données d'entrée
                if 'data' not in chart_info or not chart_info['data']:
                    logger.warning(f"Données d'image manquantes pour {chart_info.get('name', 'inconnu')}")
                    continue
                    
                # Utiliser 'data' au lieu de 'file'
                img_data = chart_info['data']
                img_buffer = io.BytesIO(img_data)
                
                # Ouvrir l'image avec Pillow
                img = Image.open(img_buffer)
                
                # Validation de l'image
                if img.width == 0 or img.height == 0:
                    logger.warning(f"Image invalide (dimensions nulles) : {chart_info.get('name', 'inconnu')}")
                    continue
                
                # Redimensionner si nécessaire (max 800x600)
                if img.width > 800 or img.height > 600:
                    img.thumbnail((800, 600), Image.Resampling.LANCZOS)
                
                # Sauvegarder en mémoire
                final_buffer = io.BytesIO()
                img.save(final_buffer, format='PNG')
                final_buffer.seek(0)
                
                # Ajouter à Excel
                openpyxl_img = OpenpyxlImage(final_buffer)
                openpyxl_img.anchor = f'A{chart_row}'
                ws.add_image(openpyxl_img)
                
                # Ajouter le nom du graphique
                title_cell = ws.cell(row=chart_row, column=6, value=chart_info.get('name', 'Graphique sans nom'))
                title_cell.font = Font(bold=True, size=12)
                
                chart_row += 25  # Espacement entre les graphiques
                chart_count += 1
                
            except Exception as e:
                logger.error(f"Erreur traitement image {chart_info.get('name', 'inconnu')}: {e}")
                # Ajouter une cellule d'erreur au lieu de planter
                error_cell = ws.cell(row=chart_row, column=1, value=f"Erreur: Impossible de charger l'image {chart_info.get('name', 'inconnu')}")
                error_cell.font = Font(color="FF0000", italic=True)
                chart_row += 5
        
        logger.info(f"Feuille Graphiques créée avec {chart_count} images")
    
    def _create_analysis_sheet(self, ws, explanations):
        """Crée la feuille d'analyses avec les explications"""
        row = 1
        for explanation in explanations:
            title_cell = ws.cell(row=row, column=1, value=explanation.get('title', ''))
            title_cell.font = Font(bold=True, size=12)
            
            content_cell = ws.cell(row=row+1, column=1, value=explanation.get('content', ''))
            content_cell.alignment = Alignment(wrap_text=True)
            row += 3
        
        # Ajuster les colonnes
        self._auto_adjust_columns(ws)
        logger.info("Feuille Analyses créée")
    
    def post(self, request):
        """Traite la demande d'export avancé avec graphiques optionnels"""
        try:
            # Récupération des métadonnées
            meta_str = request.POST.get('meta', '{}')
            metadata = json.loads(meta_str) if meta_str else {}
            
            # Récupération des données tabulaires
            tables_str = request.POST.get('tables', '[]')
            tables_data = json.loads(tables_str) if tables_str else []
            
            # Récupération des explications
            explanations_str = request.POST.get('explanations', '[]')
            explanations = json.loads(explanations_str) if explanations_str else []
            
            # Récupération des images de graphiques (optionnel)
            chart_images = []
            for key, file in request.FILES.items():
                if key.startswith('chart_'):
                    try:
                        # Traitement de l'image
                        image = Image.open(file)
                        img_buffer = io.BytesIO()
                        image.save(img_buffer, format='PNG')
                        img_buffer.seek(0)
                        
                        chart_images.append({
                            'name': file.name.replace('.png', ''),
                            'data': img_buffer.getvalue()
                        })
                    except Exception as e:
                        logger.warning(f"Erreur lors du traitement de l'image {key}: {str(e)}")
                        continue
            
            # Création du classeur Excel
            wb = Workbook()
            
            # Supprimer la feuille par défaut
            wb.remove(wb.active)
            
            # 1. Feuille Informations
            info_ws = wb.create_sheet("Informations")
            self._create_info_sheet(info_ws, metadata)
            
            # 2. Feuille Données (toujours créée)
            data_ws = wb.create_sheet("Données")
            self._create_enhanced_data_sheet(data_ws, tables_data, metadata)
            
            # 3. Feuille Graphiques (seulement si des images sont présentes)
            if chart_images:
                charts_ws = wb.create_sheet("Graphiques")
                self._create_charts_sheet(charts_ws, chart_images)
            
            # 4. Feuille Analyses (seulement si des explications sont présentes)
            if explanations:
                analysis_ws = wb.create_sheet("Analyses")
                self._create_analysis_sheet(analysis_ws, explanations)
            
            # Génération de la réponse
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            report_type = metadata.get('type', 'rapport')
            date_str = metadata.get('date', timezone.now().strftime('%Y-%m-%d'))
            filename = f"rapport_avance_{report_type}_{date_str}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            logger.info(f"Export avancé généré avec succès: {filename}")
            
            return response
            
        except json.JSONDecodeError as e:
            logger.error(f"Erreur de décodage JSON: {str(e)}")
            return JsonResponse({
                'error': 'Données JSON invalides',
                'details': str(e)
            }, status=400)
        except Exception as e:
            logger.error(f"Erreur lors de l'export avancé: {str(e)}")
            return JsonResponse({
                'error': 'Erreur lors de la génération du rapport',
                'details': str(e)
            }, status=500)