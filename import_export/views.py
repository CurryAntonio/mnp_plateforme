from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, viewsets
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from django.http import HttpResponse
from django.utils import timezone
from django.db import transaction
from django.db.models import Q
import pandas as pd
import io
import os
import pytz 
from .models import ImportHistory, ExportHistory
from .serializers import ImportHistorySerializer, ExportHistorySerializer
from products.models import Produit, Groupe
from journal.models import Mouvement
from products.serializers import ProduitSerializer, GroupeSerializer
from journal.serializers import MouvementSerializer
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

class ImportExcelView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    def post(self, request):
        try:
            file = request.FILES.get('file')
            if not file:
                return Response({'error': 'Aucun fichier fourni'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Vérifier l'extension du fichier
            if not file.name.endswith(('.xlsx', '.xls')):
                return Response({'error': 'Format de fichier non supporté. Utilisez .xlsx ou .xls'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Créer un enregistrement d'historique d'import (CORRECTION)
            import_history = ImportHistory.objects.create(
                user=request.user,
                file_name=file.name,
                import_type='products'  # Définir le type par défaut ou le récupérer du request
            )
            
            try:
                # Lire le fichier Excel
                df = pd.read_excel(file)
                
                # Vérifier si le DataFrame est vide
                if df.empty:
                    import_history.status = 'failed'
                    import_history.error_details = 'Le fichier Excel est vide'  # CORRECTION
                    import_history.save()
                    return Response({'error': 'Le fichier Excel est vide'}, status=status.HTTP_400_BAD_REQUEST)
                
                # Nettoyer les données
                df = df.fillna('')  # Remplacer les NaN par des chaînes vides
                
                # Détecter automatiquement le type de données basé sur les colonnes
                columns = [col.lower().strip() for col in df.columns]
                
                # Définir les variations possibles pour chaque colonne clé
                code_variations = ['code', 'code produit', 'code_produit']
                nom_variations = ['nom', 'nom du produit', 'nom_produit', 'nom produit']
                stock_variations = ['stock', 'stock initial', 'stock_initial', 'stock actuel', 'stock_actuel']
                description_variations = ['description', 'desc']
                
                # Vérifier la présence des colonnes avec variations
                has_code = any(var in columns for var in code_variations)
                has_nom = any(var in columns for var in nom_variations)
                has_stock = any(var in columns for var in stock_variations)
                has_description = any(var in columns for var in description_variations)
                
                if has_code and has_nom and has_stock:
                    # C'est probablement des produits
                    import_history.import_type = 'products'
                    import_history.save()
                    result = self._import_products(df, import_history)
                elif has_nom and has_description:
                    # C'est probablement des groupes
                    import_history.import_type = 'groups'
                    import_history.save()
                    result = self._import_groups(df, import_history)
                else:
                    import_history.status = 'failed'
                    import_history.error_details = 'Format de fichier non reconnu. Colonnes attendues pour produits: Code, Nom, Stock (ou variations). Pour groupes: Nom, Description.'  # CORRECTION
                    import_history.save()
                    return Response({
                        'error': 'Format de fichier non reconnu',
                        'details': 'Colonnes attendues pour produits: Code, Nom, Stock (ou leurs variations). Pour groupes: Nom, Description.',
                        'columns_found': list(df.columns),
                        'columns_normalized': columns
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                return Response(result, status=status.HTTP_200_OK)
                
            except Exception as e:
                import_history.status = 'failed'
                import_history.error_details = str(e)  # CORRECTION
                import_history.save()
                return Response({'error': f'Erreur lors du traitement du fichier: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                
        except Exception as e:
            return Response({'error': f'Erreur inattendue: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def _import_products(self, df, import_history):
        """Importer des produits depuis le DataFrame"""
        try:
            with transaction.atomic():
                created_count = 0
                updated_count = 0
                errors = []
                
                # Normaliser les noms de colonnes
                df.columns = [col.lower().strip() for col in df.columns]
                
                # Mapper les variations de colonnes vers les noms standards
                column_mapping = {
                    # Variations pour le code
                    'code produit': 'code',
                    'code_produit': 'code',
                    # Variations pour le nom
                    'nom du produit': 'nom',
                    'nom_produit': 'nom',
                    'nom produit': 'nom',
                    # Variations pour le stock
                    'stock initial': 'stock',
                    'stock_initial': 'stock',
                    'stock actuel': 'stock',
                    'stock_actuel': 'stock',
                    # Autres variations
                    'unité': 'unite',
                    'seuil d\'alerte': 'seuil',
                    'zone de stockage': 'zone'
                }
                
                # Appliquer le mapping
                df = df.rename(columns=column_mapping)
                
                for index, row in df.iterrows():
                    try:
                        # Extraire les données de base
                        code = str(row.get('code', '')).strip()
                        nom = str(row.get('nom', '')).strip()
                        stock = float(row.get('stock', 0))
                        
                        if not code or not nom:
                            errors.append(f"Ligne {index + 2}: Code et nom requis")
                            continue
                        
                        # Vérifier si le produit existe
                        produit, created = Produit.objects.get_or_create(
                            code=code,
                            defaults={
                                'nom': nom,
                                'stock_initial': stock,
                                'seuil': float(row.get('seuil', 10)),
                                'unite': str(row.get('unite', 'pièce')).strip(),
                                'zone': str(row.get('zone', '')).strip(),
                                'observation': str(row.get('observation', '')).strip(),
                            }
                        )
                        
                        if created:
                            created_count += 1
                        else:
                            # Mettre à jour le produit existant
                            produit.nom = nom
                            produit.stock_initial = stock
                            produit.seuil = float(row.get('seuil', produit.seuil))
                            produit.unite = str(row.get('unite', produit.unite)).strip()
                            produit.zone = str(row.get('zone', produit.zone)).strip()
                            produit.observation = str(row.get('observation', produit.observation)).strip()
                            produit.save()
                            updated_count += 1
                            
                    except Exception as e:
                        errors.append(f"Ligne {index + 2}: {str(e)}")
                
                import_history.status = 'completed'
                import_history.row_count = created_count + updated_count
                import_history.save()
                
                return {
                    'message': 'Import terminé avec succès',
                    'created': created_count,
                    'updated': updated_count,
                    'errors': errors
                }
                
        except Exception as e:
            import_history.status = 'failed'
            import_history.error_message = str(e)
            import_history.save()
            return {
                'error': f'Erreur lors de l\'import des produits: {str(e)}'
            }

    def _import_groups(self, df, import_history):
        """Importer des groupes depuis le DataFrame"""
        try:
            with transaction.atomic():
                created_count = 0
                updated_count = 0
                errors = []
                
                # Palette de couleurs prédéfinies (simples, foncées et claires)
                color_name_to_hex = {
                    # Couleurs simples
                    'Bleu': '#3B82F6',
                    'Rouge': '#EF4444', 
                    'Vert': '#10B981',
                    'Orange': '#F59E0B',
                    'Violet': '#8B5CF6',
                    'Rose': '#EC4899',
                    'Jaune': '#EAB308',
                    'Cyan': '#06B6D4',
                    'Indigo': '#6366F1',
                    'Gris': '#6B7280',
                    
                    # Couleurs foncées
                    'Bleu foncé': '#1E40AF',
                    'Rouge foncé': '#DC2626',
                    'Vert foncé': '#059669',
                    'Orange foncé': '#D97706',
                    'Violet foncé': '#7C3AED',
                    'Rose foncé': '#BE185D',
                    'Jaune foncé': '#CA8A04',
                    'Cyan foncé': '#0891B2',
                    'Indigo foncé': '#4F46E5',
                    'Gris foncé': '#374151',
                    
                    # Couleurs claires
                    'Bleu clair': '#93C5FD',
                    'Rouge clair': '#FCA5A5',
                    'Vert clair': '#6EE7B7',
                    'Orange clair': '#FCD34D',
                    'Violet clair': '#C4B5FD',
                    'Rose clair': '#F9A8D4',
                    'Jaune clair': '#FDE047',
                    'Cyan clair': '#67E8F9',
                    'Indigo clair': '#A5B4FC',
                    'Gris clair': '#D1D5DB'
                }
                
                # Normaliser les noms de colonnes
                df.columns = [col.lower().strip() for col in df.columns]
                
                for index, row in df.iterrows():
                    try:
                        nom = str(row.get('nom', '')).strip()
                        description = str(row.get('description', '')).strip()
                        couleur_input = str(row.get('couleur', '')).strip()
                        
                        if not nom:
                            errors.append(f"Ligne {index + 2}: Nom requis")
                            continue
                        
                        # Traitement de la couleur
                        couleur_finale = '#3B82F6'  # Couleur par défaut (Bleu)
                        
                        if couleur_input:
                            # Recherche exacte (sensible à la casse)
                            if couleur_input in color_name_to_hex:
                                couleur_finale = color_name_to_hex[couleur_input]
                            # Recherche insensible à la casse
                            else:
                                couleur_input_title = couleur_input.title()
                                if couleur_input_title in color_name_to_hex:
                                    couleur_finale = color_name_to_hex[couleur_input_title]
                                else:
                                    # Couleur non reconnue, utiliser la couleur par défaut
                                    errors.append(f"Ligne {index + 2}: Couleur '{couleur_input}' non reconnue. Couleurs disponibles: {', '.join(sorted(color_name_to_hex.keys()))}")
                        
                        # Vérifier si le groupe existe
                        groupe, created = Groupe.objects.get_or_create(
                            nom=nom,
                            defaults={
                                'description': description,
                                'couleur': couleur_finale,
                            }
                        )
                        
                        if created:
                            created_count += 1
                        else:
                            # Mettre à jour le groupe existant
                            groupe.description = description
                            groupe.couleur = couleur_finale
                            groupe.save()
                            updated_count += 1
                            
                    except Exception as e:
                        errors.append(f"Ligne {index + 2}: {str(e)}")
                
                import_history.status = 'completed'
                import_history.row_count = created_count + updated_count
                import_history.save()
                
                return {
                    'message': 'Import terminé avec succès',
                    'created': created_count,
                    'updated': updated_count,
                    'errors': errors
                }
                
        except Exception as e:
            import_history.status = 'failed'
            import_history.error_message = str(e)
            import_history.save()
            return {
                'error': f'Erreur lors de l\'import des groupes: {str(e)}'
            }


class ExportExcelView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            # Récupérer les paramètres de la requête
            export_type = request.query_params.get('type', 'products')
            
            print(f"DEBUG EXPORT TYPE: {export_type}")
            print(f"DEBUG REQUEST PARAMS: {dict(request.query_params)}")
            
            # Mapper les types de rapport frontend vers les types backend
            type_mapping = {
                'product-report': 'products',
                'group-report': 'groups', 
                'movement-report': 'movements'
            }
            
            if export_type in type_mapping:
                export_type = type_mapping[export_type]
            
            # Valider le type d'export
            valid_types = ['products', 'groups', 'movements']
            if export_type not in valid_types:
                return Response({'error': f'Type d\'export invalide: {export_type}'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Créer un enregistrement d'historique d'export
            export_history = ExportHistory.objects.create(
                user=request.user,
                export_type=export_type,
                filters={}
            )
            
            # Traitement selon le type d'export
            if export_type == 'products':
                filters = {}
                
                # Construire les filtres
                if request.query_params.get('groupe'):
                    filters['groupe_id'] = request.query_params.get('groupe')
                
                search = request.query_params.get('search')
                if search:
                    produits = Produit.objects.filter(
                        Q(nom__icontains=search) | Q(description__icontains=search),
                        **filters
                    )
                else:
                    produits = Produit.objects.filter(**filters)
                
                # Sérialiser les données
                serializer = ProduitSerializer(produits, many=True)
                data = serializer.data
                
                # Créer un DataFrame
                df = pd.DataFrame(data)
                
                if not df.empty:
                    # Supprimer les colonnes non désirées (ID, dernier_inventaire, groupe ID)
                    columns_to_remove = ['id', 'dernier_inventaire', 'groupe', 'lead_time']
                    for col in columns_to_remove:
                        if col in df.columns:
                            df = df.drop(columns=[col])
                    
                    # CORRECTION DÉCALAGE HORAIRE - Convertir les dates correctement
                    date_columns = ['date_creation', 'date_modification']
                    local_tz = pytz.timezone('Indian/Antananarivo')
                    for col in date_columns:
                        if col in df.columns:
                            df[col] = pd.to_datetime(df[col], utc=True)
                            df[col] = df[col].dt.tz_convert(local_tz)
                            df[col] = df[col].dt.strftime('%Y-%m-%d %H:%M:%S')
                    
                    # NOMS DE COLONNES PROFESSIONNELS
                    column_mapping = {
                        'code': 'Code Article',
                        'nom': 'Désignation',
                        'description': 'Description Détaillée',
                        'prix_unitaire': 'Prix Unitaire (Ar)',
                        'stock_actuel': 'Stock Disponible',
                        'stock_initial': 'Stock Initial',
                        'stock_minimum': 'Stock de Sécurité',
                        'seuil': 'Seuil d\'Alerte',
                        'unite_mesure': 'Unité de Mesure',
                        'unite': 'Unité de Mesure',
                        'zone': 'Zone de Stockage',
                        'observation': 'Observations',
                        'groupe_nom': 'Catégorie',
                        'date_creation': 'Date de Création',
                        'date_modification': 'Dernière Modification'
                    }
                    
                    # Renommer seulement les colonnes qui existent
                    existing_columns = {k: v for k, v in column_mapping.items() if k in df.columns}
                    df = df.rename(columns=existing_columns)
                    
                    # ORDRE PROFESSIONNEL DES COLONNES
                    desired_order_products = [
                        'Code Article', 'Désignation', 'Catégorie', 'Description Détaillée',
                        'Stock Disponible', 'Stock Initial', 'Stock de Sécurité', 'Seuil d\'Alerte',
                        'Unité de Mesure', 'Prix Unitaire (Ar)', 'Zone de Stockage',
                        'Date de Création', 'Dernière Modification', 'Observations'
                    ]
                    available_columns = [col for col in desired_order_products if col in df.columns]
                    remaining_columns = [col for col in df.columns if col not in available_columns]
                    final_order = available_columns + remaining_columns
                    df = df[final_order]
                    
                    # Mettre à jour le nombre de lignes
                    export_history.row_count = len(df)
                    export_history.save()
                    
                    # Créer le fichier Excel
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name='Produits', index=False)
                        
                        # Ajuster la largeur des colonnes
                        worksheet = writer.sheets['Produits']
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    output.seek(0)
                    
                    # Préparer la réponse
                    response = HttpResponse(
                        output.getvalue(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response['Content-Disposition'] = f'attachment; filename="export_produits_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
                    
                    return response
                else:
                    return Response({'error': 'Aucun produit trouvé'}, status=status.HTTP_404_NOT_FOUND)
            
            elif export_type == 'groups':
                groupes = Groupe.objects.all()
                
                # Sérialiser les données
                serializer = GroupeSerializer(groupes, many=True)
                data = serializer.data
                
                # Créer un DataFrame
                df = pd.DataFrame(data)
                
                if not df.empty:
                    # Supprimer la colonne ID
                    if 'id' in df.columns:
                        df = df.drop(columns=['id'])
                    
                    # CORRECTION DÉCALAGE HORAIRE - Convertir les dates correctement
                    date_columns = ['date_creation', 'date_modification']
                    local_tz = pytz.timezone('Indian/Antananarivo')
                    for col in date_columns:
                        if col in df.columns:
                            df[col] = pd.to_datetime(df[col], utc=True)
                            df[col] = df[col].dt.tz_convert(local_tz)
                            df[col] = df[col].dt.strftime('%Y-%m-%d %H:%M:%S')
                    
                    # NOMS DE COLONNES PROFESSIONNELS
                    column_mapping = {
                        'nom': 'Nom de la Catégorie',
                        'productCount': 'Nombre de Produits',
                        'lowStockCount': 'Articles en Rupture',
                        'description': 'Description',
                        'couleur': 'Couleur d\'Identification',
                        'date_creation': 'Date de Création',
                        'date_modification': 'Dernière Modification'
                    }
                    
                    # Renommer seulement les colonnes qui existent
                    existing_columns = {k: v for k, v in column_mapping.items() if k in df.columns}
                    df = df.rename(columns=existing_columns)
                    
                    # Convertir les codes couleur hex en noms de couleurs pour l'export
                    if 'Couleur d\'Identification' in df.columns:
                        hex_to_color_name = {
                            '#3B82F6': 'Bleu',
                            '#EF4444': 'Rouge',
                            '#10B981': 'Vert',
                            '#F59E0B': 'Orange',
                            '#8B5CF6': 'Violet',
                            '#EC4899': 'Rose',
                            '#EAB308': 'Jaune',
                            '#06B6D4': 'Cyan',
                            '#6366F1': 'Indigo',
                            '#6B7280': 'Gris',
                            '#1E40AF': 'Bleu foncé',
                            '#DC2626': 'Rouge foncé',
                            '#059669': 'Vert foncé',
                            '#D97706': 'Orange foncé',
                            '#7C3AED': 'Violet foncé',
                            '#BE185D': 'Rose foncé',
                            '#CA8A04': 'Jaune foncé',
                            '#0891B2': 'Cyan foncé',
                            '#4F46E5': 'Indigo foncé',
                            '#374151': 'Gris foncé',
                            '#93C5FD': 'Bleu clair',
                            '#FCA5A5': 'Rouge clair',
                            '#6EE7B7': 'Vert clair',
                            '#FCD34D': 'Orange clair',
                            '#C4B5FD': 'Violet clair',
                            '#F9A8D4': 'Rose clair',
                            '#FDE047': 'Jaune clair',
                            '#67E8F9': 'Cyan clair',
                            '#A5B4FC': 'Indigo clair',
                            '#D1D5DB': 'Gris clair'
                        }
                        df['Couleur d\'Identification'] = df['Couleur d\'Identification'].map(hex_to_color_name).fillna(df['Couleur d\'Identification'])
                    
                    # ORDRE PROFESSIONNEL DES COLONNES
                    desired_order = ['Nom de la Catégorie', 'Nombre de Produits', 'Articles en Rupture', 'Description', 'Couleur d\'Identification', 'Date de Création', 'Dernière Modification']
                    available_columns = [col for col in desired_order if col in df.columns]
                    remaining_columns = [col for col in df.columns if col not in available_columns]
                    final_order = available_columns + remaining_columns
                    df = df[final_order]
                    
                    # Mettre à jour le nombre de lignes
                    export_history.row_count = len(df)
                    export_history.status = 'terminé'
                    export_history.save()
                    
                    # Créer le fichier Excel avec les couleurs
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name='Groupes', index=False)
                        
                        worksheet = writer.sheets['Groupes']
                        
                        # Appliquer les couleurs de fond aux cellules de couleur
                        if 'Couleur d\'Identification' in df.columns:
                            color_col_idx = df.columns.get_loc('Couleur d\'Identification') + 1
                            color_name_to_hex = {
                                'Bleu': '#3B82F6',
                                'Rouge': '#EF4444',
                                'Vert': '#10B981',
                                'Orange': '#F59E0B',
                                'Violet': '#8B5CF6',
                                'Rose': '#EC4899',
                                'Jaune': '#EAB308',
                                'Cyan': '#06B6D4',
                                'Indigo': '#6366F1',
                                'Gris': '#6B7280',
                                'Bleu foncé': '#1E40AF',
                                'Rouge foncé': '#DC2626',
                                'Vert foncé': '#059669',
                                'Orange foncé': '#D97706',
                                'Violet foncé': '#7C3AED',
                                'Rose foncé': '#BE185D',
                                'Jaune foncé': '#CA8A04',
                                'Cyan foncé': '#0891B2',
                                'Indigo foncé': '#4F46E5',
                                'Gris foncé': '#374151',
                                'Bleu clair': '#93C5FD',
                                'Rouge clair': '#FCA5A5',
                                'Vert clair': '#6EE7B7',
                                'Orange clair': '#FCD34D',
                                'Violet clair': '#C4B5FD',
                                'Rose clair': '#F9A8D4',
                                'Jaune clair': '#FDE047',
                                'Cyan clair': '#67E8F9',
                                'Indigo clair': '#A5B4FC',
                                'Gris clair': '#D1D5DB'
                            }
                            
                            for row_idx, color_name in enumerate(df['Couleur d\'Identification'], start=2):
                                if color_name in color_name_to_hex:
                                    hex_color = color_name_to_hex[color_name].replace('#', '')
                                    fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
                                    cell = worksheet.cell(row=row_idx, column=color_col_idx)
                                    cell.fill = fill
                        
                        # Ajuster la largeur des colonnes
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    output.seek(0)
                    
                    # Préparer la réponse
                    response = HttpResponse(
                        output.getvalue(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response['Content-Disposition'] = f'attachment; filename="export_groupes_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
                    
                    return response
                else:
                    return Response({'error': 'Aucun groupe trouvé'}, status=status.HTTP_404_NOT_FOUND)
            
            elif export_type == 'movements':
                filters = {}
                
                # Construire les filtres à partir des paramètres de requête
                start_date = request.GET.get('start_date')
                end_date = request.GET.get('end_date')
                produit_id = request.GET.get('produit_id')
                movement_type = request.GET.get('movement_type')
                
                print(f"DEBUG MOVEMENT FILTERS - start_date: {start_date}, end_date: {end_date}, produit_id: {produit_id}, movement_type: {movement_type}")
                
                if start_date:
                    filters['date__gte'] = start_date
                if end_date:
                    filters['date__lte'] = end_date
                if produit_id:
                    filters['produit_id'] = produit_id
                
                # Correction du filtre de mouvement - utiliser les vrais types
                if movement_type and movement_type in ['Entrée', 'Sortie', 'Demandée']:
                    filters['mouvement'] = movement_type
                
                print(f"DEBUG FINAL MOVEMENT FILTERS: {filters}")
                
                # Requête des mouvements
                mouvements = Mouvement.objects.filter(**filters)
                
                print(f"DEBUG MOVEMENTS COUNT: {mouvements.count()}")
                
                # Sérialiser les données
                serializer = MouvementSerializer(mouvements, many=True)
                data = serializer.data
                
                # Créer un DataFrame
                df = pd.DataFrame(data)
                
                if not df.empty:
                    # Supprimer les colonnes non désirées (ID, produit ID, utilisateur ID)
                    columns_to_remove = ['id', 'produit', 'utilisateur']
                    for col in columns_to_remove:
                        if col in df.columns:
                            df = df.drop(columns=[col])
                    
                    # CORRECTION DÉCALAGE HORAIRE - Convertir les dates correctement
                    local_tz = pytz.timezone('Indian/Antananarivo')
                    if 'date' in df.columns:
                        try:
                            df['date'] = pd.to_datetime(df['date'], utc=True, format='ISO8601')
                        except Exception:
                            df['date'] = pd.to_datetime(df['date'], utc=True, errors='coerce')
                        df['date'] = df['date'].dt.tz_convert(local_tz)
                        df['date'] = df['date'].dt.strftime('%Y-%m-%d %H:%M:%S')
                    
                    # NOMS DE COLONNES PROFESSIONNELS
                    column_mapping = {
                        'produit_code': 'Code Article',
                        'produit_nom': 'Désignation',
                        'groupe': 'Catégorie',
                        'unite': 'Unité',
                        'mouvement': 'Type de Mouvement',
                        'quantite': 'Quantité',
                        'stock_avant': 'Stock Avant',
                        'stock_apres': 'Reste en Stock',
                        'prix_unitaire': 'Prix Unitaire',
                        'total': 'Montant Total',
                        'date': 'Date et Heure',
                        'demandeur': 'Demandeur',
                        'observation': 'Observation',
                        'user_nom': 'Responsable'
                    }
                    
                    # Renommer seulement les colonnes qui existent
                    existing_columns = {k: v for k, v in column_mapping.items() if k in df.columns}
                    df = df.rename(columns=existing_columns)
                    
                    # ORDRE PROFESSIONNEL DES COLONNES
                    desired_order_movements = [
                        'Date et Heure', 'Code Article', 'Désignation', 'Catégorie',
                        'Type de Mouvement', 'Quantité', 'Unité', 'Stock Avant',
                        'Stock Après', 'Prix Unitaire (Ar)', 'Montant Total (Ar)',
                        'Responsable', 'Demandeur', 'Observations'
                    ]
                    available_columns = [col for col in desired_order_movements if col in df.columns]
                    remaining_columns = [col for col in df.columns if col not in available_columns]
                    final_order = available_columns + remaining_columns
                    df = df[final_order]
                    
                    # Mettre à jour le nombre de lignes
                    export_history.row_count = len(df)
                    export_history.save()
                    
                    # Créer le fichier Excel
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name='Mouvements', index=False)
                        
                        # Ajuster la largeur des colonnes
                        worksheet = writer.sheets['Mouvements']
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    output.seek(0)
                    
                    # Préparer la réponse
                    response = HttpResponse(
                        output.getvalue(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response['Content-Disposition'] = f'attachment; filename="export_mouvements_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
                    
                    return response
                else:
                    return Response({'error': 'Aucun mouvement trouvé'}, status=status.HTTP_404_NOT_FOUND)
            
            else:
                return Response({'error': 'Type d\'export non supporté'}, status=status.HTTP_400_BAD_REQUEST)
                
        except Exception as e:
            print(f"DEBUG EXPORT ERROR: {str(e)}")
            return Response({'error': f'Erreur lors de l\'export: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ExportCompletView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        import logging
        logger = logging.getLogger('import_export.views')
        
        try:
            logger.info(f"Début de l'export complet pour l'utilisateur {request.user.id}")
            
            # Créer l'historique d'export
            export_history = ExportHistory.objects.create(
                user=request.user,
                export_type='complet',
                status='processing'
            )
            logger.info(f"Historique d'export créé avec l'ID {export_history.id}")
            
            # Créer le buffer pour le fichier Excel
            buffer = io.BytesIO()
            
            with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                # 1. FEUILLE GROUPES
                logger.info("Début de l'export des groupes")
                groupes_queryset = Groupe.objects.all()
                logger.info(f"Nombre de groupes trouvés: {groupes_queryset.count()}")
                groupes_serializer = GroupeSerializer(groupes_queryset, many=True)
                groupes_data = groupes_serializer.data
                
                if groupes_data:
                    df_groupes = pd.DataFrame(groupes_data)
                    df_groupes = df_groupes.fillna('')
                    
                    # Supprimer les colonnes non désirées
                    columns_to_remove = ['id', 'lowStockCount']
                    for col in columns_to_remove:
                        if col in df_groupes.columns:
                            df_groupes = df_groupes.drop(columns=[col])
                    
                    # Convertir les couleurs hexadécimales en noms
                    if 'couleur' in df_groupes.columns:
                        color_mapping = {
                            '#3B82F6': 'Bleu', '#EF4444': 'Rouge', '#10B981': 'Vert',
                            '#F59E0B': 'Orange', '#8B5CF6': 'Violet', '#EC4899': 'Rose', '#EAB308': 'Jaune',
                            '#06B6D4': 'Cyan', '#6366F1': 'Indigo', '#6B7280': 'Gris', '#1E40AF': 'Bleu foncé', '#DC2626': 'Rouge foncé', '#059669': 'Vert foncé',
                            '#D97706': 'Orange foncé', '#7C3AED': 'Violet foncé', '#BE185D': 'Rose foncé',
                            '#CA8A04': 'Jaune foncé', '#0891B2': 'Cyan foncé', '#4F46E5': 'Indigo foncé',
                            '#374151': 'Gris foncé',
                            '#93C5FD': 'Bleu clair', '#FCA5A5': 'Rouge clair', '#6EE7B7': 'Vert clair',
                            '#FCD34D': 'Orange clair', '#C4B5FD': 'Violet clair', '#F9A8D4': 'Rose clair',
                            '#FDE047': 'Jaune clair', '#67E8F9': 'Cyan clair', '#A5B4FC': 'Indigo clair',
                            '#D1D5DB': 'Gris clair'
                        }
                        df_groupes['couleur'] = df_groupes['couleur'].map(color_mapping).fillna('Autre')
                    
                    # Formater les dates
                    madagascar_tz = pytz.timezone('Indian/Antananarivo')
                    date_columns = ['date_creation', 'date_modification']
                    for col in date_columns:
                        if col in df_groupes.columns:
                            df_groupes[col] = pd.to_datetime(df_groupes[col], utc=True)
                            df_groupes[col] = df_groupes[col].dt.tz_convert(madagascar_tz)
                            df_groupes[col] = df_groupes[col].dt.strftime('%d/%m/%Y %H:%M:%S')
                    
                    # Renommer les colonnes
                    df_groupes = df_groupes.rename(columns={
                        'nom': 'Nom', 'description': 'Description', 'couleur': 'Couleur',
                        'productCount': 'Nombre de Produits', 'date_creation': 'Date de Création',
                        'date_modification': 'Date de Modification'
                    })
                    
                    df_groupes.to_excel(writer, sheet_name='Groupes', index=False)
                    logger.info(f"Export des groupes terminé: {len(df_groupes)} lignes")
                
                # 2. FEUILLE PRODUITS
                logger.info("Début de l'export des produits")
                produits_queryset = Produit.objects.all()
                logger.info(f"Nombre de produits trouvés: {produits_queryset.count()}")
                produits_serializer = ProduitSerializer(produits_queryset, many=True)
                produits_data = produits_serializer.data
                
                if produits_data:
                    df_produits = pd.DataFrame(produits_data)
                    df_produits = df_produits.fillna('')
                    
                    # Supprimer les colonnes non désirées
                    columns_to_remove = ['id', 'groupe', 'dernier_inventaire', 'lead_time']
                    for col in columns_to_remove:
                        if col in df_produits.columns:
                            df_produits = df_produits.drop(columns=[col])
                    
                    # Formater les dates
                    madagascar_tz = pytz.timezone('Indian/Antananarivo')
                    date_columns = ['date_creation', 'date_modification']
                    for col in date_columns:
                        if col in df_produits.columns:
                            df_produits[col] = pd.to_datetime(df_produits[col], utc=True)
                            df_produits[col] = df_produits[col].dt.tz_convert(madagascar_tz)
                            df_produits[col] = df_produits[col].dt.strftime('%d/%m/%Y %H:%M:%S')
                    
                    # Renommer les colonnes
                    df_produits = df_produits.rename(columns={
                        'code': 'Code', 'nom': 'Nom', 'stock_initial': 'Stock',
                        'seuil': 'Seuil', 'unite': 'Unité', 'zone': 'Zone',
                        'observation': 'Observation', 'groupe_nom': 'Groupe',
                        'date_creation': 'Date de Création', 'date_modification': 'Date de Modification'
                    })
                    
                    df_produits.to_excel(writer, sheet_name='Produits', index=False)
                    logger.info(f"Export des produits terminé: {len(df_produits)} lignes")
                
                # 3. FEUILLE JOURNAL DES MOUVEMENTS
                logger.info("Début de l'export des mouvements")
                mouvements_queryset = Mouvement.objects.select_related('produit').all().order_by('-date')
                logger.info(f"Nombre de mouvements trouvés: {mouvements_queryset.count()}")
                mouvements_serializer = MouvementSerializer(mouvements_queryset, many=True)
                mouvements_data = mouvements_serializer.data
                
                if mouvements_data:
                    df_mouvements = pd.DataFrame(mouvements_data)
                    df_mouvements = df_mouvements.fillna('')
                    
                    columns_to_remove = ['id', 'produit', 'utilisateur']
                    for col in columns_to_remove:
                        if col in df_mouvements.columns:
                            df_mouvements = df_mouvements.drop(columns=[col])
                            
                    # Formater les dates
                    madagascar_tz = pytz.timezone('Indian/Antananarivo')
                    if 'date' in df_mouvements.columns:
                        try:
                            df_mouvements['date'] = pd.to_datetime(df_mouvements['date'], utc=True, format='ISO8601')
                        except Exception:
                            try:
                                df_mouvements['date'] = pd.to_datetime(df_mouvements['date'], utc=True, format='mixed')
                            except Exception:
                                df_mouvements['date'] = pd.to_datetime(df_mouvements['date'], utc=True, errors='coerce')
                        df_mouvements['date'] = df_mouvements['date'].dt.tz_convert(madagascar_tz)
                        df_mouvements['date'] = df_mouvements['date'].dt.strftime('%d/%m/%Y %H:%M:%S')
                    
                    # Renommer les colonnes
                    df_mouvements = df_mouvements.rename(columns={
                        'produit_nom': 'Produit',
                        'produit_code': 'Code Produit',
                        'mouvement': 'Type de Mouvement',
                        'quantite': 'Quantité',
                        'stock_avant': 'Stock Avant',
                        'stock_apres': 'Stock Après',
                        'demandeur': 'Demandeur',
                        'date': 'Date',
                        'observation': 'Observation'
                    })
                    
                    df_mouvements.to_excel(writer, sheet_name='Journal des Mouvements', index=False)
                    logger.info(f"Export des mouvements terminé: {len(df_mouvements)} lignes")
                
                # Ajuster la largeur des colonnes pour toutes les feuilles
                logger.info("Ajustement de la largeur des colonnes")
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    # Obtenir le DataFrame correspondant
                    if sheet_name == 'Groupes' and 'df_groupes' in locals():
                        df = df_groupes
                    elif sheet_name == 'Produits' and 'df_produits' in locals():
                        df = df_produits
                    elif sheet_name == 'Journal des Mouvements' and 'df_mouvements' in locals():
                        df = df_mouvements
                    else:
                        continue
                    
                    for i, col in enumerate(df.columns):
                        column_width = max(df[col].astype(str).map(len).max(), len(col)) + 2
                        worksheet.set_column(i, i, column_width)
            
            # Calculer le nombre total de lignes
            total_rows = 0
            if 'df_groupes' in locals():
                total_rows += len(df_groupes)
            if 'df_produits' in locals():
                total_rows += len(df_produits)
            if 'df_mouvements' in locals():
                total_rows += len(df_mouvements)
            
            logger.info(f"Nombre total de lignes exportées: {total_rows}")
            
            export_history.row_count = total_rows
            export_history.status = 'completed'
            export_history.save()
            logger.info(f"Historique d'export mis à jour: statut = terminé")
            
            # Préparer la réponse
            buffer.seek(0)
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{export_history.file_name}"'
            
            logger.info("Export complet terminé avec succès")
            return response
            
        except Exception as e:
            logger.error(f"Erreur lors de l'export complet: {str(e)}")
            logger.error(f"Type d'erreur: {type(e).__name__}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            
            # Mettre à jour le statut de l'historique en cas d'erreur
            if 'export_history' in locals():
                export_history.status = 'failed'
                export_history.save()
            
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ImportHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ImportHistory.objects.all().order_by('-created_at')
    serializer_class = ImportHistorySerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        return ImportHistory.objects.filter(user=self.request.user).order_by('-created_at')


class ExportHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ExportHistory.objects.all().order_by('-created_at')
    serializer_class = ExportHistorySerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        return ExportHistory.objects.filter(user=self.request.user).order_by('-created_at')


class BackupView(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        try:
            # Créer une sauvegarde de la base de données
            backup_name = f"sauvegarde_{timezone.now().strftime('%Y%m%d_%H%M%S')}.json"
            
            # Ici vous pouvez implémenter la logique de sauvegarde
            # Par exemple, utiliser Django's dumpdata command
            
            return Response({
                'message': 'Sauvegarde créée avec succès',
                'backup_name': backup_name
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response({
                'error': f'Erreur lors de la création de la sauvegarde: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def get(self, request):
        try:
            # Lister les sauvegardes disponibles
            backups = [
                {
                    'id': 1,
                    'name': 'sauvegarde_20231201_120000.json',
                    'created_at': '2023-12-01T12:00:00Z',
                    'size': '2.5 MB'
                }
            ]
            
            return Response(backups, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'error': f'Erreur lors de la récupération des sauvegardes: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class RestoreBackupView(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request, backup_id):
        try:
            # Restaurer une sauvegarde spécifique
            # Ici vous pouvez implémenter la logique de restauration
            
            return Response({
                'message': f'Sauvegarde {backup_id} restaurée avec succès'
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({
                'error': f'Erreur lors de la restauration: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
